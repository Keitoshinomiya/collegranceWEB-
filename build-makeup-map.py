#!/usr/bin/env python3
"""
products.json ↔ メイクアップExcel 商品コードマッピング生成

一度だけ実行し、makeup_code_map.json を生成する。
以後の在庫更新は商品コードで完全一致するため誤マッチなし。

使い方:
  python3 build-makeup-map.py <Excel.xlsx>
  → makeup_code_map.json 生成
  → マッチ結果を表示（要確認リストも出力）
"""
from __future__ import annotations
import openpyxl, json, math, re, sys, os

PRODUCTS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'products.json')
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'makeup_code_map.json')

# ブランド名 日本語→英語
BRAND_MAP = {
    'ドルチェ＆ガッバーナ': 'Dolce & Gabbana', 'モスキーノ': 'Moschino',
    'ディプティック': 'DIPTYQUE', 'バーバリー': 'BURBERRY',
    'カルバンクライン': 'Calvin Klein', 'クロエ': 'CHLOE',
    'ディオール': 'DIOR', 'ジバンシイ': 'GIVENCHY', 'グッチ': 'GUCCI',
    'エルメス': 'HERMES', 'イッセイミヤケ': 'ISSEY MIYAKE',
    'ジョーマローン': 'Jo Malone London', 'ランコム': 'LANCOME',
    'メゾンマルジェラ': 'Maison Margiela', 'プラダ': 'PRADA',
    'サンローラン': 'YSL', 'トムフォード': 'TOM FORD',
    'ヴェルサーチェ': 'VERSACE', 'ブルガリ': 'BVLGARI',
    'ジミーチュウ': 'Jimmy Choo', 'コーチ': 'COACH', 'ロエベ': 'LOEWE',
    'バイレード': 'BYREDO', 'ジョルジオアルマーニ': 'Giorgio Armani',
    'マークジェイコブス': 'MARC JACOBS', 'フェラガモ': 'FERRAGAMO',
    'クリード': 'CREED', 'モンブラン': 'MONTBLANC', 'クリーン': 'CLEAN',
    'ティファニー': 'TIFFANY & CO.', 'ナルシソロドリゲス': 'NARCISO RODRIGUEZ',
    'ミュウミュウ': 'MIU MIU',
}

# 商品名 日本語→英語キーワード（Excel商品名 → products.jsonのname/nameJaにマッチさせる用）
NAME_MAP = {
    # D&G
    'ザ　ワン': 'The One', 'ザ　ワン　フォーメン': 'The One for Men',
    'NEW ライトブルー': 'Light Blue', 'ライトブルー': 'Light Blue',
    'ライトブルー　サマー　バイブス': 'Light Blue Summer Vibes',
    'ドルチェ＆ガッバーナ ライトブルー カプリインラブ': 'Capri in Love',
    'ドルチェ＆ガッバーナ ライトブルー プールオム カプリインラブ': 'Capri in Love Pour Homme',
    'ライトブルー　プールオム　2025': 'プールオム 2025',
    'ドルチェ リリー': 'Dolce Lily', 'ドルチェ バイオレット': 'Dolce Violet',
    'ドルチェ ブルージャスミン': 'Blue Jasmine',
    'ライトブルー　ラブイズラブ': 'Love is Love',
    'ライトブルー　2025　テスター': '2025 テスター',
    '旧 ライトブルー': '旧 ライトブルー',
    # VERSACE
    'ヴェルサーチェ　プールオム': 'Pour Homme',
    'ヴェルサーチェ　マン　オーフレッシュ': 'Man Eau Fra',
    'ブライト　クリスタル': 'Bright Crystal', 'ブライトクリスタル': 'Bright Crystal',
    'ヴェルセンス': 'Versense', 'イエローダイアモンド': 'Yellow Diamond',
    'エロス': 'Eros', 'エロス  フェム': 'エロス フェム', 'エロス　フェム': 'エロス フェム',
    'ディランブルー': 'Dylan Blue', 'ディランブルー　フェム': 'Dylan Blue Pour Femme',
    'エロス　フレイム': 'Eros Flame', 'ディランターコイズ': 'Dylan Turquoise',
    'ディランパープル　フェム': 'Dylan Purple',
    # Calvin Klein
    'シーケーワン': 'ck one', 'シーケー　ワン': 'ck one',
    'エタニティ': 'Eternity', 'エタニティ　フォーメン': 'Eternity for Men',
    'シーケービー': 'CK Be', 'シーケー　ワン　ゴールド': 'ck one Gold',
    'シーケー　オール': 'CK All', 'シーケー　エブリワン': 'ck Everyone',
    'シーケー　ワン　エッセンス　インテンス　パルファム': 'エッセンス インテンス',
    'エタニティ　フォーウィメン　インテンス': 'エタニティ フォーウィメン インテンス',
    'デファイ': 'デファイ',
    'エタニティフォーウィメン　アンバーエッセンス　P　インテンス': 'アンバーエッセンス',
    'エタニティフォーメン　アンバーエッセンス　P　インテンス': 'エタニティフォーメン アンバー',
    'エタニティ　フォー　メン　アロマティック　エッセンス　パルファム　インテンス': 'アロマティック',
    # LOEWE
    'ロエベ　001　ウーマン【EDT】': '001 Woman EDT',
    'ロエベ　001　マン【EDT】': '001 Man EDT',
    'ロエベ　001　ウーマン【EDP】': '001 Woman EDP',
    'ロエベ　001　マン【EDP】': '001 Man EDP',
    # テスター版
    '【テスター】　ロエベ　001　ウーマン　【EDT】': 'ウーマン 【EDT】 Tester',
    '【テスター】　ロエベ　001　マン　【EDT】': 'マン 【EDT】 Tester',
    '【テスター】　ロエベ　001　ウーマン　【EDP】': 'ウーマン 【EDP】 Tester',
    # DIPTYQUE
    'オルフェオン': 'Orpheon', 'フルール　ドゥ　ポー': 'Fleur de Peau',
    'オーローズ': 'Eau Rose', 'タム　ダオ': 'Tam Dao',
    'フィロシコス': 'Philosykos', 'ド　ソン': 'Do Son', 'オーデサンス': 'Eau des Sens',
    # BYREDO
    'ブランシュ': 'Blanche', 'ジプシー　ウォーター': 'Gypsy Water',
    # Jo Malone
    'イングリッシュペアー＆フリージア': 'English Pear',
    'ブラックベリー＆ベイ': 'Blackberry',
    'アールグレー＆キューカンバー': 'Earl Grey',
    'ネクタリンブロッサム＆ハニー': 'Nectarine',
    # BVLGARI
    'ブルガリ　プールオム': 'Pour Homme', 'オムニア　アメジスト': 'Omnia Amethyste',
    'オムニア　クリスタリン': 'Omnia Crystalline',
    'ブルガリ　プールオム　【EDPタイプ】': 'Pour Homme EDP',
    # Armani
    'アクアディジオプールオム': 'Acqua di Gio',
    'アクアディジョイア': 'Acqua di Gioia',
    # ISSEY MIYAKE
    'ロードゥイッセイ': "L'Eau d'Issey",
    'ロードゥイッセイ　プールオム': "Pour Homme",
    'ロードゥイッセイ　プールオム　ウッド＆ウッド': 'Wood & Wood',
    'ロードゥイッセイ　ピオニー　オードトワレインテンス': 'Pivoine Intense',
    # COACH
    'コーチ　【EDT】': 'コーチ 【EDT】', 'コーチ　【EDP】': 'コーチ 【EDP】',
    'コーチ　マン': 'コーチ マン', 'コーチ　ラブ': 'コーチ ラブ',
    'コーチ　ドリームス　サンセット': 'ドリームス サンセット',
    'コーチ　ドリームス　ムーンライト': 'ドリームス ムーンライト',
    'コーチ　マン　【EDP】': 'マン 【EP】', 'コーチ　ゴールド': 'コーチ ゴールド',
    # Jimmy Choo
    'ジミーチュウ': 'Jimmy Choo EDP',
    'ジミーチュウ　マン': 'ジミーチュウ マン',
    'ジミーチュウ　ロー': 'ジミーチュウ ロー',
    'ジミーチュウ　マンアイス': 'マンアイス',
    'ジミーチュウ　マンブルー': 'マンブルー',
    'ジミーチュウ　フローラル': 'フローラル',
    'ジミー チュウ アーバンヒーロー': 'アーバンヒーロー',
    'アイ・ウォント・チュウ': 'アイ・ウォント・チュウ',
    'ジミーチュウ　マン　アクア': 'マン アクア',
    'ジミー チュウ ローズ パッション': 'ローズ パッション',
    'アイ・ウォント・チュウ　ル　パルファム': 'ル パルファム',
    'ジミーチュウ　マン　エクストリーム': 'エクストリーム',
    # CLEAN
    'クラシック　ウォームコットン': 'ウォームコットン',
    'クラシック　フレッシュランドリー': 'フレッシュランドリー',
    'クラシック　ピュアソープ': 'ピュアソープ',
    'クラシック　アップルブロッサム': 'アップルブロッサム',
    'クラシック　スプリングブリーズ': 'スプリング ブリーズ',
    'クラシック　ストロベリーフィールド': 'ストロベリーフィールド',
    'クラシック　シャワーフレッシュ': 'Shower Fresh',
    'クラシック　フレッシュリネン': 'フレッシュリネン',
    'クラシック　クールコットン': 'クールコットン',
    'クラシック　フラワーフレッシュ': 'フラワーフレッシュ',
    # Moschino
    'モスキーノ・トイ2　パール': 'Toy 2 Pearl',
    'モスキーノ・トイ・ボーイ': 'Toy Boy',
    'モスキーノ・トイ2 バブルガム': 'Toy 2 Bubblegum',
    # BURBERRY
    'マイ　バーバリー　ブラッシュ': 'Burberry Blush',
    '【テスター品】　ウィークエンド　フォーメン': 'Weekend for Men Tester',
    '【テスター品】　バーバリータッチ　フォーメン': 'Touch for Men Tester',
    # TOM FORD
    'ネロリ　ポルトフィーノ': 'Neroli Portofino',
    'ブラック　オーキッド': 'Black Orchid',
    'ビター　ピーチ': 'Bitter Peach',
    'タバコ・バニラ': 'Tobacco Vanille',
    'ロストチェリー': 'Lost Cherry',
    # GUCCI
    'ラッシュ': 'Rush',
    # PRADA
    'キャンディ': 'Candy', 'パラドックス': 'Paradoxe',
    # TIFFANY
    'ローズゴールド': 'Rose Gold',
}

def extract_ml(s):
    m = re.search(r'(\d+)\s*m[lL]', s.replace('ｍ','m').replace('Ｌ','L'))
    return int(m.group(1)) if m else None

def load_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if str(row[14] or '') != '香水':
            continue
        items.append({
            'code': str(row[2] or '').strip(),
            'brand_jp': str(row[4] or '').strip(),
            'name_jp': str(row[5] or '').strip(),
            'spec': str(row[6] or '').strip(),
            'size': str(row[7] or '').strip(),
            'stock': str(row[11] or ''),
            'cost': row[13],
        })
    return items

def match_product_to_excel(product, excel_items):
    """products.jsonの1商品 → Excel商品コードをマッチ"""
    p_brand = product['brand'].upper()
    p_ml = extract_ml(product.get('size', ''))
    p_name = (product.get('name', '') + ' ' + product.get('nameJa', '')).strip()
    is_tester = 'tester' in p_name.lower() or 'テスター' in p_name

    candidates = []
    for ei in excel_items:
        brand_en = BRAND_MAP.get(ei['brand_jp'])
        if not brand_en or brand_en.upper() != p_brand:
            continue
        e_ml = extract_ml(ei['size'])
        if e_ml != p_ml:
            continue

        # テスター判定
        e_tester = 'テスター' in ei['name_jp']
        if is_tester != e_tester:
            continue

        # 名前マッチスコア
        mapped_name = NAME_MAP.get(ei['name_jp'], '')
        if not mapped_name:
            # テスター除去して再検索
            clean_name = ei['name_jp'].replace('【テスター】　', '').replace('【テスター品】　', '')
            mapped_name = NAME_MAP.get(clean_name, '')

        score = 0
        if mapped_name:
            # mapped_nameのキーワードがp_nameに含まれるか
            for kw in mapped_name.split():
                if kw.lower() in p_name.lower():
                    score += 10

        candidates.append({'excel': ei, 'score': score, 'mapped_name': mapped_name})

    if not candidates:
        return None, 'no_candidate'

    # スコア順ソート
    candidates.sort(key=lambda x: x['score'], reverse=True)

    if len(candidates) == 1:
        return candidates[0]['excel']['code'], 'unique'

    # 最高スコアが1つだけなら確定
    top = candidates[0]['score']
    top_candidates = [c for c in candidates if c['score'] == top]

    if len(top_candidates) == 1 and top > 0:
        return top_candidates[0]['excel']['code'], 'scored'

    # 曖昧
    return candidates, 'ambiguous'

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 build-makeup-map.py <Excel.xlsx>")
        sys.exit(1)

    excel_path = sys.argv[1]
    excel_items = load_excel(excel_path)
    print(f"Excel香水: {len(excel_items)}件")

    with open(PRODUCTS_JSON) as f:
        products = json.load(f)
    print(f"products.json: {len(products)}件")
    print()

    code_map = {}  # product_id -> makeup_code
    auto_matched = []
    ambiguous_list = []
    no_match = []

    for p in products:
        result, status = match_product_to_excel(p, excel_items)

        if status == 'unique' or status == 'scored':
            code_map[str(p['id'])] = result
            auto_matched.append({
                'id': p['id'], 'brand': p['brand'], 'name': p['name'],
                'size': p.get('size',''), 'code': result, 'status': status
            })
        elif status == 'ambiguous':
            ambiguous_list.append({
                'id': p['id'], 'brand': p['brand'], 'name': p['name'],
                'size': p.get('size',''), 'candidates': result
            })
        else:
            no_match.append(p)

    # 結果表示
    print(f"✅ 自動マッチ: {len(auto_matched)}件")
    print(f"⚠️ 要確認（候補複数）: {len(ambiguous_list)}件")
    print(f"❌ 候補なし（メイクアップ外）: {len(no_match)}件")
    print()

    print("=== 自動マッチ結果 ===")
    for m in auto_matched:
        print(f"  [{m['id']:>3}] {m['brand']} {m['name']} ({m['size']}) → {m['code']}")

    print()
    print("=== 要確認（候補複数） ===")
    for item in ambiguous_list:
        print(f"\n  [{item['id']:>3}] {item['brand']} {item['name']} ({item['size']})")
        for c in item['candidates'][:5]:
            ei = c['excel']
            print(f"       {ei['code']} | {ei['name_jp']} | {ei['spec']} | {ei['size']} | score={c['score']}")

    print()
    print("=== 候補なし（メイクアップ外の仕入れ） ===")
    for p in no_match:
        print(f"  [{p['id']:>3}] {p['brand']} {p['name']} ({p.get('size','')})")

    # 保存
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(code_map, f, indent=2, ensure_ascii=False)
    print(f"\n→ {OUTPUT_FILE} に{len(code_map)}件のマッピングを保存しました")


if __name__ == '__main__':
    main()
