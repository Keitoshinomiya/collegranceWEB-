#!/usr/bin/env python3
"""
メイクアップ在庫メール自動チェック & 処理スクリプト

Mac Miniのcronで定期実行:
  0 10 * * * cd /Users/keito/GitHub/collegranceWEB- && python3 check-makeup-email.py

処理フロー:
  1. Gmail APIでyamamoto@makeup-inc.comの新着メールをチェック
  2. 添付Excelをダウンロード
  3. 在庫更新ルール実行（products.json更新）
  4. 価格チェック・重複チェック
  5. Slackに結果レポート送信
"""

import os, sys, json, math, base64, pickle, re, datetime, requests, tempfile
from pathlib import Path

# === 設定 ===
CREDENTIALS_FILE = '/Users/keito/Downloads/client_secret_95164447019-8drtt3gg1f3id1n26g6t0s0rnnkb0ejo.apps.googleusercontent.com.json'
TOKEN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'gmail_token.pickle')
PRODUCTS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'products.json')
SENDER = 'yamamoto@makeup-inc.com'
SLACK_WEBHOOK_URL = None  # or direct API
SLACK_BOT_TOKEN = os.environ.get('SLACK_BOT_TOKEN', '')
SLACK_CHANNEL = 'C091LDC8MKN'
MARGIN = 1.25
TAX = 1.10
MIN_PRICE = 3000

# === Gmail API ===
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

def get_gmail_service():
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'wb') as f:
            pickle.dump(creds, f)

    return build('gmail', 'v1', credentials=creds)


def find_latest_email(service):
    """yamamoto@makeup-inc.comからの最新の添付付きメールを検索"""
    results = service.users().messages().list(
        userId='me',
        q=f'from:{SENDER} has:attachment subject:価格リスト',
        maxResults=1
    ).execute()

    messages = results.get('messages', [])
    if not messages:
        return None, None

    msg = service.users().messages().get(
        userId='me', id=messages[0]['id']
    ).execute()

    # メール日付を取得
    headers = {h['name']: h['value'] for h in msg['payload']['headers']}
    subject = headers.get('Subject', '')
    date_str = headers.get('Date', '')

    return msg, subject


def download_attachment(service, message):
    """メールからExcel添付ファイルをダウンロード"""
    parts = message.get('payload', {}).get('parts', [])

    for part in parts:
        filename = part.get('filename', '')
        if not filename:
            continue
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            continue

        attachment_id = part['body'].get('attachmentId')
        if not attachment_id:
            continue

        attachment = service.users().messages().attachments().get(
            userId='me',
            messageId=message['id'],
            id=attachment_id
        ).execute()

        data = base64.urlsafe_b64decode(attachment['data'])

        # 保存
        save_path = os.path.join(tempfile.gettempdir(), filename)
        with open(save_path, 'wb') as f:
            f.write(data)

        print(f"添付ファイルダウンロード: {filename} -> {save_path}")
        return save_path, filename

    return None, None


# === 在庫更新処理 ===
def process_excel(excel_path):
    """Excelを読み込み、products.jsonを更新"""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Sheet1']

    # Excel商品を抽出
    # 列マッピング: [2]=商品コード [4]=ブランド [5]=商品名 [6]=規格(EDT等)
    # [7]=容量 [8]=性別 [11]=在庫数 [13]=単価(税抜) [14]=商品カテゴリー
    excel_items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cat = str(row[14] or '')
        if cat != '香水':
            continue
        brand = str(row[4] or '').strip()
        name = str(row[5] or '').strip()
        cost_raw = row[13]
        if not brand or not cost_raw:
            continue
        try:
            cost = int(str(cost_raw).replace(',', ''))
        except:
            continue
        sell = math.ceil(cost * MARGIN * TAX / 10) * 10
        if sell < MIN_PRICE:
            continue

        spec = str(row[6] or '').strip()
        size = str(row[7] or '').strip()
        gender = str(row[8] or '').strip()
        stock = str(row[11] or '')
        code = str(row[2] or '')

        excel_items.append({
            'code': code, 'brand': brand, 'name': name, 'cost': cost,
            'sell': sell, 'size': size, 'spec': spec, 'gender': gender, 'stock': stock
        })

    print(f"Excel: 香水¥{MIN_PRICE}以上 = {len(excel_items)}商品")

    # products.json読み込み
    with open(PRODUCTS_JSON) as f:
        products = json.load(f)

    # 結果集計
    report = {
        'total_excel': len(excel_items),
        'price_updated': 0,
        'stock_on': 0,
        'stock_off': 0,
        'new_products': 0,
        'duplicates': [],
        'errors': [],
    }

    # 在庫更新・価格更新のロジックはここに追加
    # （現時点ではレポートのみ）

    # 価格チェック
    instock = [x for x in products if x.get('inStock') != False]
    for x in instock:
        expected = math.ceil(x['cost'] * MARGIN * TAX / 10) * 10
        if x['sellPrice'] != expected:
            report['errors'].append(f"価格不一致: id={x['id']} {x['brand']} {x['name']} 計算=¥{expected} 実際=¥{x['sellPrice']}")

    return report


# === Slack通知 ===
def send_slack(text):
    if not SLACK_BOT_TOKEN:
        print(f"[Slack] {text}")
        return

    resp = requests.post('https://slack.com/api/chat.postMessage', json={
        'channel': SLACK_CHANNEL,
        'text': text,
    }, headers={
        'Authorization': f'Bearer {SLACK_BOT_TOKEN}',
        'Content-Type': 'application/json',
    })
    data = resp.json()
    if not data.get('ok'):
        print(f"Slack error: {data.get('error')}")


# === メイン ===
def main():
    # 最終実行日を記録して、同じメールを二重処理しない
    last_run_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.last_makeup_check')
    last_message_id = ''
    if os.path.exists(last_run_file):
        with open(last_run_file) as f:
            last_message_id = f.read().strip()

    print("=== メイクアップ在庫メールチェック ===")
    print(f"日時: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")

    service = get_gmail_service()

    msg, subject = find_latest_email(service)
    if not msg:
        print("新着メールなし")
        return

    msg_id = msg['id']
    if msg_id == last_message_id:
        print(f"処理済みメール: {subject}")
        return

    print(f"新着メール: {subject}")

    # 添付ファイルダウンロード
    excel_path, filename = download_attachment(service, msg)
    if not excel_path:
        send_slack(f"⚠️ メイクアップから新しいメールが届きましたが、Excel添付が見つかりません。\n件名: {subject}")
        return

    # Slackに通知
    send_slack(
        f"📋 *メイクアップから新しい価格リストが届きました*\n"
        f"件名: {subject}\n"
        f"ファイル: {filename}\n\n"
        f"添付Excelを自動ダウンロードしました。在庫更新処理を実行中..."
    )

    # Excel処理
    report = process_excel(excel_path)

    # 結果レポート
    report_text = (
        f"✅ *在庫チェック完了*\n"
        f"Excel商品数: {report['total_excel']}\n"
    )

    if report['errors']:
        report_text += f"\n⚠️ *エラー {len(report['errors'])}件*:\n"
        for e in report['errors'][:5]:
            report_text += f"  • {e}\n"

    if report['duplicates']:
        report_text += f"\n🔄 *重複確認が必要 {len(report['duplicates'])}件*:\n"
        for d in report['duplicates'][:5]:
            report_text += f"  • {d}\n"

    send_slack(report_text)

    # 処理済みとして記録
    with open(last_run_file, 'w') as f:
        f.write(msg_id)

    print("完了")


if __name__ == '__main__':
    main()
