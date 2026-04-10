#!/usr/bin/env python3
"""
商品価格の完全チェック — 在庫更新後に必ず実行
  python check-prices.py
  python check-prices.py --verbose  (詳細表示)

チェック項目:
  1. 計算チェック: sellPrice = cost × 1.25 × 1.10 (10円切上)
  2. サイズ整合: 大サイズが小サイズより安くないか
  3. 価格範囲: ¥1,000未満 or ¥60,000超は異常
  4. 濃度漏れ: 全商品にconcentration必須
  5. 重複: 同ブランド・同名・同サイズ・同濃度で複数存在しないか
  6. メイクアップExcel照合: products.jsonのcostがExcelと一致しているか
  7. k-styleスプシ照合: k-styleの卸値と比較し、高い方を採用しているか
  8. 画像漏れ: 在庫あり商品に画像パスが設定されているか
"""
import json, math, re, os, sys
from collections import defaultdict

MARGIN = 1.25
TAX = 1.10

with open('products.json') as f:
    products = json.load(f)

instock = [x for x in products if x.get('inStock') != False]
errors = []
warnings = []
verbose = '--verbose' in sys.argv

# ============================================================
# CHECK 1: sellPrice = cost × 1.25 × 1.10 (10円切上)
# ============================================================
for x in instock:
    expected = math.ceil(x['cost'] * MARGIN * TAX / 10) * 10
    if x['sellPrice'] != expected:
        errors.append(f"[計算] id={x['id']} {x['brand']} {x['name']} cost=¥{x['cost']:,} 計算=¥{expected:,} 実際=¥{x['sellPrice']:,}")

# ============================================================
# CHECK 2: 同一商品の大サイズが小サイズより安くないか
# ============================================================
groups = defaultdict(list)
for x in instock:
    base = x['name'].split(' (')[0].split(' 【')[0].replace('NEW','').replace('旧 ','').strip()
    groups[(x['brand'], base)].append(x)

for key, items in groups.items():
    if len(items) < 2: continue
    def get_ml(item):
        m = re.search(r'(\d+)', item['size'])
        return int(m.group(1)) if m else 0
    sorted_items = sorted(items, key=get_ml)
    for i in range(len(sorted_items) - 1):
        a, b = sorted_items[i], sorted_items[i+1]
        if get_ml(b) > get_ml(a) > 0 and b['sellPrice'] < a['sellPrice']:
            errors.append(f"[サイズ逆転] {a['brand']} {a['name']} {a['size']}=¥{a['sellPrice']:,} > {b['size']}=¥{b['sellPrice']:,}")

# ============================================================
# CHECK 3: 異常価格
# ============================================================
for x in instock:
    if x['sellPrice'] < 1000:
        errors.append(f"[安すぎ] id={x['id']} {x['brand']} {x['name']} ¥{x['sellPrice']:,}")
    if x['sellPrice'] > 60000:
        errors.append(f"[高すぎ] id={x['id']} {x['brand']} {x['name']} ¥{x['sellPrice']:,}")

# ============================================================
# CHECK 4: 濃度漏れ
# ============================================================
for x in instock:
    if not x.get('concentration'):
        errors.append(f"[濃度なし] id={x['id']} {x['brand']} {x['name']}")

# ============================================================
# CHECK 5: 重複（同ブランド・同名・同サイズ・同濃度）
# ============================================================
for key, items in groups.items():
    for i in range(len(items)):
        for j in range(i+1, len(items)):
            a, b = items[i], items[j]
            if a['size'].lower().replace('ml','') == b['size'].lower().replace('ml','') and a.get('concentration','') == b.get('concentration','') and a['sellPrice'] != b['sellPrice']:
                errors.append(f"[同サイズ同濃度で価格不一致] id={a['id']} vs id={b['id']} {a['brand']} {a['name'][:20]} {a['size']} ¥{a['sellPrice']:,} vs ¥{b['sellPrice']:,}")

# ============================================================
# CHECK 6: メイクアップExcel照合
# ============================================================
excel_checked = 0
excel_path = None
for path in [
    '/var/folders/j2/0zgl7dr134x85pl1swjd9_zw0000gn/T/2026_4_6　価格リスト.xlsx',
    '/Users/keito/Downloads/2026_4_6　価格リスト.xlsx',
    '/Users/keito/Downloads/2026_4_1　価格リスト.xlsx',
]:
    if os.path.exists(path):
        excel_path = path
        break

if excel_path:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Sheet1']

        brand_map = {
            'HERMÈS': 'エルメス', 'Maison Margiela': 'メゾン　マルジェラ', 'BYREDO': 'バイレード',
            'Calvin Klein': 'カルバンクライン', 'Dolce & Gabbana': 'ドルチェ＆ガッバーナ',
            'TIFFANY & CO.': 'ティファニー', 'ISSEY MIYAKE': 'イッセイミヤケ',
            'YSL': 'イヴサンローラン', 'LOEWE': 'ロエベ', 'CLEAN': 'クリーン',
            'Jo Malone London': 'ジョーマローン', 'DIPTYQUE': 'ディプティック',
            'Chloé': 'クロエ', 'VERSACE': 'ヴェルサーチェ', 'Jimmy Choo': 'ジミーチュウ',
            'COACH': 'コーチ', 'LANVIN': 'ランバン', 'BVLGARI': 'ブルガリ',
            'Giorgio Armani': 'ジョルジオアルマーニ', 'BURBERRY': 'バーバリー',
            'Anna Sui': 'アナスイ', 'Lancôme': 'ランコム', 'TOM FORD': 'トムフォード',
            'Salvatore Ferragamo': 'サルヴァトーレフェラガモ', 'GUCCI': 'グッチ',
            'PRADA': 'プラダ', 'GIVENCHY': 'ジバンシイ', 'Clinique': 'クリニーク',
            'MOSCHINO': 'モスキーノ', 'Kate Spade': 'ケイト・スペード・ニューヨーク',
            'Mercedes-Benz': 'メルセデスベンツ', 'NISHANE': 'ニシャネ',
        }

        excel_items = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            cat = str(row[13] or '')
            brand = str(row[4] or '').strip()
            name = str(row[5] or '').strip()
            cost = row[12]
            size = str(row[7] or '').strip()
            if not brand or not cost: continue
            try:
                cost = int(str(cost).replace(',',''))
            except: continue
            excel_items.append({'brand': brand, 'name': name, 'cost': cost, 'size': size})

        for p in instock:
            brand_ja = brand_map.get(p['brand'], '')
            if not brand_ja: continue
            p_size = p['size'].lower().replace('ml','').strip()
            p_name = p['name'].lower()

            for e in excel_items:
                if brand_ja not in e['brand']: continue
                e_size = e['size'].lower().replace('ml','').strip()
                if e_size != p_size: continue
                e_name = e['name'].lower().replace('　',' ')
                p_words = [w for w in p_name.split() if len(w) > 2][:2]
                if not any(w in e_name for w in p_words): continue

                excel_checked += 1
                # products.jsonのcostがExcel以上であることを確認（高い方ルール）
                if p['cost'] < e['cost']:
                    warnings.append(f"[Excel卸値より安い] id={p['id']} {p['brand']} {p['name']} {p['size']}: products=¥{p['cost']:,} < Excel=¥{e['cost']:,}")
                break

        if verbose:
            print(f"  Excel照合: {excel_checked}商品をチェック ({excel_path})")
    except ImportError:
        warnings.append("[Excel照合スキップ] openpyxlがインストールされていません")
else:
    warnings.append("[Excel照合スキップ] Excelファイルが見つかりません")

# ============================================================
# CHECK 7: k-styleスプシ照合
# ============================================================
kstyle_data = {
    ('DIPTYQUE', 'Orpheon', '75'): 20500,
    ('DIPTYQUE', 'Fleur de Peau', '75'): 20500,
    ('YSL', 'MYSLF', '100'): 13500,
    ('YSL', 'Mon Paris', '90'): 13800,
    ('HERMÈS', 'ナイルの庭', '100'): 8600,
    ('HERMÈS', '地中海の庭', '100'): 10900,
    ('HERMÈS', '李氏の庭', '100'): 10050,
    ('Chloé', 'Chloé EDP', '75'): 8500,
    ('Jo Malone London', 'Wood Sage', '100'): 11500,
    ('Jo Malone London', 'Blackberry', '100'): 11500,
    ('Jo Malone London', 'English Pear', '100'): 11500,
    ('DIOR', 'Sauvage', '200'): 20000,
    ('TIFFANY & CO.', 'Rose Gold', '75'): 11200,
    ('BYREDO', 'Blanche', '100'): 27000,
    ('Maison Margiela', 'Lazy Sunday', '100'): 9000,
    ('LE LABO', 'Another 13', '100'): 37000,
    ('LE LABO', 'Santal 33', '100'): 34000,
    ('Maison Margiela', 'Coffee Break', '100'): 10500,
    ('Maison Margiela', 'Jazz Club', '100'): 9800,
    ('BVLGARI', 'Omnia Crystalline', '100'): 10100,
    ('BYREDO', 'Gypsy Water', '100'): 23000,
    ('Calvin Klein', 'ck one', '200'): 3150,
    ('DIOR', 'Hypnotic Poison', '150'): 20500,
    ('YSL', 'LIBRE', '90'): 15250,
    ('DIPTYQUE', 'Tam Dao', '100'): 16500,
    ('DECORTE', 'Kimono Yui', '50'): 6300,
}

kstyle_checked = 0
for (brand, name_part, size), k_cost in kstyle_data.items():
    for p in instock:
        if p['brand'] != brand: continue
        p_size = re.search(r'(\d+)', p['size'])
        if not p_size or p_size.group(1) != size: continue
        if name_part.lower() not in p['name'].lower() and name_part.lower() not in (p.get('nameJa') or '').lower(): continue

        kstyle_checked += 1
        # products.jsonのcostがk-style以上であることを確認（高い方ルール）
        if p['cost'] < k_cost:
            errors.append(f"[k-style卸値より安い] id={p['id']} {p['brand']} {p['name']} {p['size']}: products=¥{p['cost']:,} < k-style=¥{k_cost:,}")
        elif p['cost'] > k_cost:
            # 高い方を使っているのでOK（メイクアップの方が高い）
            if verbose:
                print(f"  OK: id={p['id']} {p['brand']} {p['name']}: products=¥{p['cost']:,} > k-style=¥{k_cost:,} (高い方採用)")
        break

if verbose:
    print(f"  k-style照合: {kstyle_checked}商品をチェック")

# ============================================================
# CHECK 8: 画像漏れ
# ============================================================
for x in instock:
    if not x.get('img'):
        errors.append(f"[画像パスなし] id={x['id']} {x['brand']} {x['name']}")
    elif not os.path.exists(x['img']):
        errors.append(f"[画像ファイルなし] id={x['id']} {x['brand']} {x['name']} ({x['img']})")

# ============================================================
# RESULT
# ============================================================
print(f"=== 価格チェック（{len(instock)}商品）===\n")

if errors:
    print(f"❌ エラー: {len(errors)}件\n")
    for e in errors:
        print(f"  {e}")

if warnings:
    print(f"\n⚠ 警告: {len(warnings)}件\n")
    for w in warnings:
        print(f"  {w}")

if not errors and not warnings:
    print(f"  ✅ 全チェック合格")
    print(f"     計算・サイズ整合・価格範囲・濃度・重複・Excel照合({excel_checked})・k-style照合({kstyle_checked})・画像")
elif not errors:
    print(f"\n  ✅ エラーなし（警告{len(warnings)}件は確認推奨）")
