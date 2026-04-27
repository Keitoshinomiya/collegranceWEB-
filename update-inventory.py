#!/usr/bin/env python3
"""
メイクアップExcel → products.json 在庫・価格更新スクリプト

makeup_code_map.json のマッピングを使い、商品コードで完全一致。
誤マッチなし。

使い方:
  python3 update-inventory.py <Excel.xlsx>           # ドライラン（変更内容を表示）
  python3 update-inventory.py <Excel.xlsx> --apply    # 実際にproducts.jsonを更新
"""
from __future__ import annotations
import openpyxl, json, math, re, sys, os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCTS_JSON = os.path.join(SCRIPT_DIR, 'products.json')
CODE_MAP_FILE = os.path.join(SCRIPT_DIR, 'makeup_code_map.json')
MARGIN = 1.25
TAX = 1.10

def load_excel(path):
    """Excelを読み込み、商品コード→データのdictを返す"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    items = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if str(row[14] or '') != '香水':
            continue
        code = str(row[2] or '').strip()
        if not code:
            continue
        stock_raw = str(row[11] or '')
        in_stock = False
        if '以上' in stock_raw:
            in_stock = True
        else:
            try:
                in_stock = int(stock_raw) > 0
            except:
                in_stock = False
        cost = None
        if row[13]:
            try:
                cost = int(str(row[13]).replace(',', ''))
            except:
                pass
        items[code] = {
            'brand': str(row[4] or '').strip(),
            'name': str(row[5] or '').strip(),
            'spec': str(row[6] or '').strip(),
            'size': str(row[7] or '').strip(),
            'stock': stock_raw,
            'in_stock': in_stock,
            'cost': cost,
        }
    return items

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 update-inventory.py <Excel.xlsx> [--apply]")
        sys.exit(1)

    excel_path = sys.argv[1]
    apply = '--apply' in sys.argv

    # マッピング読み込み
    with open(CODE_MAP_FILE) as f:
        code_map = json.load(f)  # {"product_id": "makeup_code"}

    # Excel読み込み
    excel_data = load_excel(excel_path)
    print(f"Excel: {len(excel_data)}件の香水商品")

    # products.json読み込み
    with open(PRODUCTS_JSON) as f:
        products = json.load(f)
    print(f"products.json: {len(products)}件")
    print(f"マッピング: {len(code_map)}件")
    print()

    # 変更リスト
    stock_on = []   # 在庫なし → あり
    stock_off = []  # 在庫あり → なし
    cost_changed = []  # 卸値変更
    not_in_excel = []  # マッピングあるがExcelにコードなし

    for p in products:
        pid = str(p['id'])
        makeup_code = code_map.get(pid)
        if not makeup_code:
            continue  # メイクアップ外の商品

        excel_item = excel_data.get(makeup_code)
        if not excel_item:
            not_in_excel.append({
                'id': p['id'], 'brand': p['brand'], 'name': p['name'],
                'code': makeup_code
            })
            continue

        was_in = p.get('inStock', True) != False
        now_in = excel_item['in_stock']

        # 在庫変更
        if not was_in and now_in:
            stock_on.append({
                'id': p['id'], 'brand': p['brand'], 'name': p['name'],
                'size': p.get('size', ''), 'stock': excel_item['stock'],
                'excel_name': excel_item['name'], 'code': makeup_code,
            })
        elif was_in and not now_in:
            stock_off.append({
                'id': p['id'], 'brand': p['brand'], 'name': p['name'],
                'size': p.get('size', ''), 'stock': excel_item['stock'],
                'excel_name': excel_item['name'], 'code': makeup_code,
            })

        # 卸値変更
        if excel_item['cost'] and p.get('cost'):
            if excel_item['cost'] != p['cost']:
                new_sell = math.ceil(excel_item['cost'] * MARGIN * TAX / 10) * 10
                cost_changed.append({
                    'id': p['id'], 'brand': p['brand'], 'name': p['name'],
                    'size': p.get('size', ''),
                    'old_cost': p['cost'], 'new_cost': excel_item['cost'],
                    'old_sell': p['sellPrice'], 'new_sell': new_sell,
                    'excel_name': excel_item['name'], 'code': makeup_code,
                })

    # レポート
    print("=" * 60)
    print(f"📦 在庫復活（なし→あり）: {len(stock_on)}件")
    for s in stock_on:
        print(f"  ✅ [{s['id']:>3}] {s['brand']} {s['name']} ({s['size']}) 在庫:{s['stock']} ← {s['code']}:{s['excel_name']}")

    print()
    print(f"⛔ 在庫切れ（あり→なし）: {len(stock_off)}件")
    for s in stock_off:
        print(f"  ❌ [{s['id']:>3}] {s['brand']} {s['name']} ({s['size']}) ← {s['code']}:{s['excel_name']}")

    print()
    print(f"💰 卸値変更: {len(cost_changed)}件")
    for c in cost_changed:
        diff = c['new_cost'] - c['old_cost']
        arrow = '↑' if diff > 0 else '↓'
        print(f"  {arrow} [{c['id']:>3}] {c['brand']} {c['name']} ({c['size']}) "
              f"¥{c['old_cost']}→¥{c['new_cost']} (販売¥{c['old_sell']}→¥{c['new_sell']}) "
              f"← {c['code']}:{c['excel_name']}")

    if not_in_excel:
        print()
        print(f"⚠️ Excelにコードなし: {len(not_in_excel)}件")
        for n in not_in_excel:
            print(f"  [{n['id']:>3}] {n['brand']} {n['name']} code={n['code']}")

    # 適用
    if apply:
        print()
        print("=" * 60)
        print("🔄 products.json に変更を適用中...")

        changes = 0
        for p in products:
            pid = str(p['id'])
            makeup_code = code_map.get(pid)
            if not makeup_code:
                continue
            excel_item = excel_data.get(makeup_code)
            if not excel_item:
                continue

            # 在庫更新
            now_in = excel_item['in_stock']
            was_in = p.get('inStock', True) != False
            if was_in != now_in:
                if now_in:
                    if 'inStock' in p:
                        del p['inStock']  # デフォルトtrue
                else:
                    p['inStock'] = False
                changes += 1

            # 卸値・販売価格更新
            if excel_item['cost'] and p.get('cost') and excel_item['cost'] != p['cost']:
                p['cost'] = excel_item['cost']
                p['sellPrice'] = math.ceil(excel_item['cost'] * MARGIN * TAX / 10) * 10
                changes += 1

        with open(PRODUCTS_JSON, 'w', encoding='utf-8') as f:
            json.dump(products, f, indent=2, ensure_ascii=False)

        print(f"✅ {changes}件の変更を適用しました")
    else:
        print()
        print("=" * 60)
        print("ℹ️ ドライラン（変更は適用されていません）")
        print("  実際に適用するには: python3 update-inventory.py <Excel> --apply")


if __name__ == '__main__':
    main()
