#!/usr/bin/env python3
"""
卸値 → 販売価格 自動更新スクリプト

卸業者の価格リスト（Excel）または納入履歴スプレッドシートから
最新の卸値を取得し、products.json の販売価格を更新する。

計算式: 販売価格 = 卸値（税抜）× 1.20（マージン）× 1.10（消費税10%）
        10円単位で切り上げ

使い方:
  python update-prices.py                      # 卸リストから更新
  python update-prices.py --source wholesale   # 卸リストから更新（明示）
  python update-prices.py --margin 1.25        # マージン25%に変更
  python update-prices.py --dry-run            # 変更をプレビューのみ
"""

import json
import math
import argparse
import os
import re
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCTS_JSON = os.path.join(SCRIPT_DIR, "products.json")

# 卸業者の価格リストのパス
WHOLESALE_XLSX = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-shinomiya.keito@onlinestore-bio.com/"
    "マイドライブ/2025_9_3　価格リスト  合同会社ヤシノミ様.xlsx"
)

DEFAULT_MARKUP = 1.20
TAX_RATE = 1.10


def calc_price(cost, markup=DEFAULT_MARKUP):
    """販売価格を計算（10円単位切り上げ）"""
    return math.ceil(cost * markup * TAX_RATE / 10) * 10


def load_wholesale_prices(xlsx_path):
    """卸業者のExcelから商品コード → 上代のマッピングを取得"""
    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl が必要です")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    prices = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        code = str(row[2].value or "").strip()
        brand = str(row[4].value or "")
        name = str(row[5].value or "")
        size = str(row[7].value or "")
        price = row[9].value
        stock = row[10].value

        if code and price:
            try:
                price = int(float(str(price).replace(",", "")))
            except (ValueError, TypeError):
                continue
            try:
                stock = int(float(str(stock).replace(",", ""))) if stock else 0
            except (ValueError, TypeError):
                stock = 0

            prices[code] = {
                "brand": brand,
                "name": name,
                "size": size,
                "cost": price,
                "stock": stock,
            }

    return prices


def update_products(markup=DEFAULT_MARKUP, dry_run=False):
    """products.json の価格を更新"""
    with open(PRODUCTS_JSON, "r", encoding="utf-8") as f:
        products = json.load(f)

    # 卸リストを読み込み
    wholesale = {}
    if os.path.exists(WHOLESALE_XLSX):
        wholesale = load_wholesale_prices(WHOLESALE_XLSX)
        print(f"卸リスト読み込み: {len(wholesale)}商品")
    else:
        print(f"卸リストが見つかりません: {WHOLESALE_XLSX}")

    changes = []
    for p in products:
        old_cost = p["cost"]
        old_price = p.get("sellPrice", calc_price(old_cost, markup))
        new_cost = old_cost
        source = "unchanged"

        # 卸リストからの商品（wholesaleCode）で更新
        wcode = p.get("wholesaleCode", "")
        if wcode and wcode in wholesale:
            new_cost = wholesale[wcode]["cost"]
            source = f"wholesale ({wcode})"

        new_price = calc_price(new_cost, markup)

        if new_cost != old_cost or new_price != old_price:
            changes.append({
                "sku": p.get("sku", "?"),
                "brand": p["brand"],
                "name": p["name"],
                "old_cost": old_cost,
                "new_cost": new_cost,
                "old_price": old_price,
                "new_price": new_price,
                "source": source,
            })

        p["cost"] = new_cost
        p["sellPrice"] = new_price

    # 変更レポート
    if changes:
        print(f"\n価格変更: {len(changes)}件")
        print(f"{'SKU':<20} {'商品名':<25} {'旧卸値':>8} {'新卸値':>8} {'旧価格':>8} {'新価格':>8}")
        print("-" * 100)
        for c in changes:
            print(f"{c['sku']:<20} {c['name']:<25} ¥{c['old_cost']:>6,} ¥{c['new_cost']:>6,} ¥{c['old_price']:>6,} ¥{c['new_price']:>6,}")
    else:
        print("\n価格変更なし")

    if not dry_run:
        with open(PRODUCTS_JSON, "w", encoding="utf-8") as f:
            json.dump(products, f, ensure_ascii=False, indent=2)
        print(f"\n✅ {PRODUCTS_JSON} を更新しました")
    else:
        print(f"\n(dry-run: 実際のファイルは変更していません)")

    return products


def generate_js(products):
    """products.json から members-catalog.html 用のJSコードを生成"""
    lines = ["const PRODUCTS = ["]
    for p in products:
        img = p.get("img", "")
        tags_str = ",".join(f"'{t}'" for t in p.get("tags", []))
        lines.append(
            f"  {{ id:{p['id']}, brand:'{p['brand']}', name:'{p['name']}', "
            f"nameJa:'{p.get('nameJa','')}', size:'{p['size']}', "
            f"cost:{p['cost']}, img:'{img}', "
            f"notes:'{p.get('notes','')}', popular:{str(p.get('popular',False)).lower()}, "
            f"tags:[{tags_str}], stripePriceId:'{p.get('stripePriceId','')}' }},"
        )
    lines.append("];")
    return "\n".join(lines)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="卸値→販売価格 自動更新")
    parser.add_argument("--margin", type=float, default=DEFAULT_MARKUP, help="マージン率（デフォルト1.20=20%%）")
    parser.add_argument("--dry-run", action="store_true", help="変更をプレビューのみ")
    parser.add_argument("--generate-js", action="store_true", help="JS用コードを出力")
    args = parser.parse_args()

    products = update_products(markup=args.margin, dry_run=args.dry_run)

    if args.generate_js:
        print("\n" + "=" * 60)
        print(generate_js(products))
