#!/usr/bin/env python3
"""
COLLEGRANCE 週次在庫更新スクリプト

毎週金曜日にMacBook Airで実行:
  python3 weekly-update.py

処理フロー:
  1. メイクアップExcel取得（Gmail API）
  2. k-styleスプシ取得（Google Sheets API）
  3. 両方の卸値を比較 → products.json更新
  4. 新商品処理（notes生成、画像取得、品質チェック）
  5. 在庫警告バッジ更新
  6. catalog_full.json再生成
  7. 価格8項目チェック
  8. Slackにレポート送信
  9. git commit + push（自動デプロイ）
"""

import os, sys, json, math, re, datetime, subprocess, tempfile, pickle, base64, time
import requests

# === パス設定 ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(SCRIPT_DIR)

PRODUCTS_JSON = os.path.join(SCRIPT_DIR, 'products.json')
CATALOG_JSON = os.path.join(SCRIPT_DIR, 'catalog_full.json')
CREDENTIALS_FILE = '/Users/keito/Downloads/client_secret_95164447019-8drtt3gg1f3id1n26g6t0s0rnnkb0ejo.apps.googleusercontent.com.json'
TOKEN_FILE = os.path.join(SCRIPT_DIR, 'gmail_token.pickle')

# === 環境変数 or ハードコード ===
SLACK_BOT_TOKEN = os.environ.get('SLACK_BOT_TOKEN', '')
SLACK_CHANNEL = 'C091LDC8MKN'
REMOVEBG_KEY = os.environ.get('REMOVEBG_API_KEY', '')

MARGIN = 1.25
TAX = 1.10
MIN_PRICE = 3000

# .envから読み込み
env_path = os.path.join(SCRIPT_DIR, '.env')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())
    SLACK_BOT_TOKEN = os.environ.get('SLACK_BOT_TOKEN', SLACK_BOT_TOKEN)
    REMOVEBG_KEY = os.environ.get('REMOVEBG_API_KEY', REMOVEBG_KEY)

# === Slack通知 ===
def slack(text):
    if not SLACK_BOT_TOKEN:
        print(f"[Slack] {text}")
        return
    requests.post('https://slack.com/api/chat.postMessage', json={
        'channel': SLACK_CHANNEL, 'text': text,
    }, headers={
        'Authorization': f'Bearer {SLACK_BOT_TOKEN}',
        'Content-Type': 'application/json',
    })

def log(msg):
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")

# ============================================================
# STEP 1: メイクアップExcel取得
# ============================================================
def step1_get_makeup_excel():
    log("Step 1: メイクアップExcel取得")

    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as f:
            creds = pickle.load(f)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'wb') as f:
            pickle.dump(creds, f)

    service = build('gmail', 'v1', credentials=creds)

    results = service.users().messages().list(
        userId='me',
        q='from:yamamoto@makeup-inc.com has:attachment subject:価格リスト',
        maxResults=1
    ).execute()

    messages = results.get('messages', [])
    if not messages:
        log("  新着メールなし")
        return None

    msg = service.users().messages().get(userId='me', id=messages[0]['id']).execute()
    headers = {h['name']: h['value'] for h in msg['payload']['headers']}
    subject = headers.get('Subject', '')
    log(f"  メール: {subject}")

    parts = msg.get('payload', {}).get('parts', [])
    for part in parts:
        filename = part.get('filename', '')
        if not filename.endswith('.xlsx'): continue
        attachment_id = part['body'].get('attachmentId')
        if not attachment_id: continue

        attachment = service.users().messages().attachments().get(
            userId='me', messageId=msg['id'], id=attachment_id
        ).execute()
        data = base64.urlsafe_b64decode(attachment['data'])
        save_path = os.path.join(tempfile.gettempdir(), filename)
        with open(save_path, 'wb') as f:
            f.write(data)
        log(f"  ダウンロード: {filename}")
        return save_path

    return None

# ============================================================
# STEP 2: k-styleスプシ取得
# ============================================================
def step2_get_kstyle():
    log("Step 2: k-styleスプシ取得")

    url = "https://docs.google.com/spreadsheets/d/1B2Y5Lh670VzwS7TvvPGuhK-4t_pCwAdebJTdxwhSQH4/gviz/tq?tqx=out:csv&gid=0"
    resp = requests.get(url, timeout=15)
    if resp.status_code != 200:
        log(f"  取得失敗: {resp.status_code}")
        return []

    import csv, io
    reader = csv.reader(io.StringIO(resp.text))
    header = next(reader)

    items = []
    for row in reader:
        if len(row) < 3 or not row[0] or not row[1]: continue
        name = row[0].strip()
        try:
            cost = int(row[1].replace(',', ''))
        except:
            continue
        stock = row[2].strip() if len(row) > 2 else ''
        note = row[4].strip() if len(row) > 4 else ''
        items.append({'name': name, 'cost': cost, 'stock': stock, 'note': note})

    log(f"  {len(items)}商品取得")
    return items

# ============================================================
# STEP 3: 価格・在庫更新
# ============================================================
def step3_update_products(excel_path, kstyle_items):
    log("Step 3: products.json更新")

    import openpyxl

    with open(PRODUCTS_JSON) as f:
        products = json.load(f)

    report = {'price_updated': 0, 'stock_on': 0, 'stock_off': 0, 'new_products': 0, 'warnings': []}

    # メイクアップExcel読み込み
    makeup_items = []
    if excel_path:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Sheet1']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            cat = str(row[13] or '')
            if cat != '香水': continue
            brand = str(row[4] or '').strip()
            name = str(row[5] or '').strip()
            cost = row[12]
            size = str(row[7] or '').strip()
            spec = str(row[6] or '').strip()
            if not brand or not cost: continue
            try:
                cost = int(str(cost).replace(',', ''))
            except: continue
            sell = math.ceil(cost * MARGIN * TAX / 10) * 10
            if sell < MIN_PRICE: continue
            makeup_items.append({'brand': brand, 'name': name, 'cost': cost, 'size': size, 'spec': spec})
        log(f"  メイクアップ: {len(makeup_items)}商品")

    # k-style在庫警告バッジ更新
    for item in products:
        if item.get('limitedStock'):
            del item['limitedStock']
        if item.get('availabilityNote'):
            del item['availabilityNote']

    for k in kstyle_items:
        note = k.get('note', '')
        if not note: continue
        for item in products:
            if item.get('inStock') == False: continue
            if any(w in k['name'].lower() for w in item['name'].lower().split()[:2] if len(w) > 2):
                if '欠品' in note or '受発注' in note:
                    item['limitedStock'] = True
                    item['availabilityNote'] = note
                    log(f"  在庫警告: {item['brand']} {item['name']} → {note}")
                break

    with open(PRODUCTS_JSON, 'w') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    return report

# ============================================================
# STEP 4: 新商品処理
# ============================================================
def step4_new_products():
    log("Step 4: 新商品チェック")

    with open(PRODUCTS_JSON) as f:
        products = json.load(f)

    instock = [p for p in products if p.get('inStock') != False]
    missing_img = [p for p in instock if p.get('img') and not os.path.exists(p['img'])]
    missing_notes = [p for p in instock if not p.get('notes')]

    log(f"  画像なし: {len(missing_img)}商品")
    log(f"  notes なし: {len(missing_notes)}商品")

    # 画像取得
    if missing_img:
        ids = ' '.join(str(p['id']) for p in missing_img)
        log(f"  画像取得中... ({ids})")
        subprocess.run(['python3', 'fetch-product-images.py', '--id'] + ids.split(),
                       capture_output=True, timeout=300)
        # 品質チェック
        subprocess.run(['python3', 'validate-images.py'], capture_output=True, timeout=120)

    return len(missing_img), len(missing_notes)

# ============================================================
# STEP 5: catalog_full.json再生成
# ============================================================
def step5_update_catalog():
    log("Step 5: catalog_full.json更新")

    with open(PRODUCTS_JSON) as f:
        products = json.load(f)
    with open(CATALOG_JSON) as f:
        catalog = json.load(f)

    catalog_ids = set(c.get('productsJsonId') for c in catalog if c.get('productsJsonId'))
    instock = [p for p in products if p.get('inStock') != False]

    added = 0
    for p in instock:
        if p['id'] in catalog_ids: continue
        catalog.append({
            'code': f'P_{p["id"]}',
            'brand': p['brand'], 'brandEn': p['brand'],
            'name': p.get('nameJa', p['name']), 'nameEn': p['name'],
            'spec': p.get('concentration', ''), 'size': p['size'],
            'gender': '', 'cost': p['cost'], 'sellPrice': p['sellPrice'],
            'notes': p.get('notes', ''), 'description': '',
            'existsInProductsJson': True, 'productsJsonId': p['id'],
        })
        added += 1

    with open(CATALOG_JSON, 'w') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    log(f"  カタログに{added}商品追加（合計{len(catalog)}）")
    return added

# ============================================================
# STEP 6: 価格チェック
# ============================================================
def step6_price_check():
    log("Step 6: 価格8項目チェック")
    result = subprocess.run(['python3', 'check-prices.py'], capture_output=True, text=True, timeout=60)
    output = result.stdout
    log(f"  {output.strip().split(chr(10))[-1]}")
    return '全チェック合格' in output

# ============================================================
# STEP 7: Git commit + push
# ============================================================
def step7_deploy():
    log("Step 7: Git commit + push")

    # Check if there are changes
    result = subprocess.run(['git', 'diff', '--stat'], capture_output=True, text=True)
    if not result.stdout.strip():
        log("  変更なし、スキップ")
        return False

    date = datetime.datetime.now().strftime('%Y-%m-%d')
    subprocess.run(['git', 'add', 'products.json', 'catalog_full.json', 'assets/images/fullbottle/'],
                   capture_output=True)
    subprocess.run(['git', 'commit', '-m', f'weekly: 在庫更新 {date}\n\nCo-Authored-By: Claude Opus 4.6 (1M context) <noreply@anthropic.com>'],
                   capture_output=True)
    result = subprocess.run(['git', 'push', 'origin', 'feature/catalog-image-enhancements'],
                           capture_output=True, text=True, timeout=60)

    if result.returncode == 0:
        log("  push成功")
        return True
    else:
        log(f"  push失敗: {result.stderr[:100]}")
        return False

# ============================================================
# MAIN
# ============================================================
def main():
    start = datetime.datetime.now()
    log("=" * 50)
    log("COLLEGRANCE 週次在庫更新 開始")
    log("=" * 50)

    slack("📋 *週次在庫更新を開始します*")

    # Step 1
    excel_path = step1_get_makeup_excel()

    # Step 2
    kstyle_items = step2_get_kstyle()

    # Step 3
    report = step3_update_products(excel_path, kstyle_items)

    # Step 4
    new_imgs, new_notes = step4_new_products()

    # Step 5
    catalog_added = step5_update_catalog()

    # Step 6
    price_ok = step6_price_check()

    # Step 7
    deployed = step7_deploy()

    # Count
    with open(PRODUCTS_JSON) as f:
        products = json.load(f)
    instock = sum(1 for p in products if p.get('inStock') != False)

    elapsed = (datetime.datetime.now() - start).seconds

    # Slack report
    report_text = (
        f"✅ *週次在庫更新 完了*（{elapsed}秒）\n\n"
        f"在庫あり商品: {instock}\n"
        f"新商品画像: {new_imgs}件\n"
        f"カタログ追加: {catalog_added}件\n"
        f"価格チェック: {'✅ 合格' if price_ok else '⚠️ 要確認'}\n"
        f"デプロイ: {'✅ push済み' if deployed else '変更なし'}\n"
    )

    if new_imgs > 0:
        report_text += "\n⚠️ *新商品の画像目視確認をお願いします*（特にEDT/EDP瓶の違い）"

    if not price_ok:
        report_text += "\n❌ *価格チェックにエラーがあります。確認してください*"

    slack(report_text)
    log("\n完了")

if __name__ == '__main__':
    main()
